"""
Generate test spreadsheets for the Spreadsheet sub-agent
(POST /api/v1/agent-conversations/sub-agents/spreadsheet).

Produces three .xlsx files that exercise the LLM column-mapping path:
  1. staff_standard.xlsx  - plain English headers (Name/Department/Position)
  2. staff_unspecified.xlsx - reordered + oddly-named headers, extra columns
  3. staff_korean.xlsx     - Korean headers (이름/부서/직급)

Each file uses the same people seeded by seed_staff_department.sql, plus two
non-existent names to exercise the `unmatched` path.

Run:  python src/test/data/gen_spreadsheets.py
"""
import os
from openpyxl import Workbook

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# (name, department, position) — first 12 match the seed SQL; last 2 are unmatched.
PEOPLE = [
    ("John Doe", "Sales", "Manager"),
    ("Mike Ross", "Sales", "Staff"),
    ("Rachel Zane", "Sales", "Associate"),
    ("Jane Smith", "Marketing", "Specialist"),
    ("Tom Hardy", "Marketing", "Lead"),
    ("Alice Johnson", "Engineering", "Senior Engineer"),
    ("Bob Martin", "Engineering", "Engineer"),
    ("David Kim", "Finance", "Analyst"),
    ("Grace Lee", "Human Resources", "HR Manager"),
    ("Henry Cho", "Human Resources", "Recruiter"),
    ("Nonexistent Person", "Sales", "Ghost"),     # unmatched
    ("Imaginary Staff", "Marketing", "Phantom"),  # unmatched
]


def save(wb, name):
    path = os.path.join(OUT_DIR, name)
    wb.save(path)
    print("wrote", path)


def standard():
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff"
    ws.append(["Name", "Department", "Position"])
    for name, dept, pos in PEOPLE:
        ws.append([name, dept, pos])
    save(wb, "staff_standard.xlsx")


def unspecified():
    # Reordered columns, non-obvious header names, plus an extra unrelated column.
    wb = Workbook()
    ws = wb.active
    ws.title = "List"
    ws.append(["No.", "Role", "Full Name", "Email", "Org Unit"])
    for i, (name, dept, pos) in enumerate(PEOPLE, start=1):
        email = name.lower().replace(" ", ".") + "@testcorp.com"
        ws.append([i, pos, name, email, dept])
    save(wb, "staff_unspecified.xlsx")


def korean():
    wb = Workbook()
    ws = wb.active
    ws.title = "직원목록"
    ws.append(["이름", "부서", "직급"])
    for name, dept, pos in PEOPLE:
        ws.append([name, dept, pos])
    save(wb, "staff_korean.xlsx")


if __name__ == "__main__":
    standard()
    unspecified()
    korean()
    print("done")
